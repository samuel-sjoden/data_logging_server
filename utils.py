import time
import re
from openpyxl import load_workbook

# worksheets in feeding rate workbook
input_worksheet = "Mass Flow Input"
output_worksheet = "Mass Flow Output"
datalog_worksheet = "Datalog"
output_and_yield_worksheet = "Output and Yield"

# Patterns for common datalog observations
hopper_smoke_start_pattern = r"Hopper Smoke Start"
hopper_smoke_finish_pattern = r"Hopper Smoke Finish"
chimney_smoke_start_pattern = r"Chimney Smoke Start"
chimney_smoke_finish_pattern = r"Chimney Smoke Finish"

# Cell locations inside of worksheets
input_bucket_cell = 15
output_bucket_cell = 15
input_moisture_cell = 43
output_moisture_cell = 54
water_cell = 32
datalog_cell = 7

# Mass of input bucket including tele-handler bucket (kg)
input_bucket_mass = 520
# Flow rate in L/min. Default is 1L/min
water_flow_rate = 1
# Start and Finish time as a Float
water_start_time = 8.0
water_finish_time = 16.5


def handle_input(server, client_addr, workbook_path):
    """
    Logs an input bucket event with a timestamp and predefined mass into the workbook.
    Sends a confirmation message to the client.
    :param (socket.socket) server: The server which the client is listening to
    :param (tuple) client_addr: The IP and port number of the client
    :param (str) workbook_path: Path to the feeding log workbook
    :return:
    """
    wb = load_workbook(workbook_path)
    global input_bucket_cell
    timestamp = time.strftime("%H:%M")
    ws = wb[input_worksheet]
    ws[f'B{input_bucket_cell}'] = timestamp
    ws[f'C{input_bucket_cell}'] = input_bucket_mass
    input_bucket_cell += 1
    wb.save(workbook_path)
    print("Before sending response...")
    server.sendto(f"Logged Input Bucket At: {timestamp}".ljust(100, '\0').encode(), client_addr)
    print("Response sent!")


def handle_output(server, client_addr, workbook_path):
    """
    Requests and logs an output bucket mass reading with a timestamp.
    Sends a confirmation message to the client.
    :param (socket.socket) server: The server which the client is listening to
    :param (tuple) client_addr: The IP and port number of the client
    :param (str) workbook_path: Path to the feeding log workbook
    """
    global output_bucket_cell
    wb = load_workbook(workbook_path)
    timestamp = time.strftime("%H:%M")
    ws = wb[output_worksheet]
    mass = get_data(server, f"Enter Measured Mass:".ljust(100, '\0'), client_addr, r"^\d{2,4}(\.\d)?$",
                    f"Invalid Entry. Please only enter digits in number of kg or send 'c' to cancel".ljust(100, '\0'))

    ws[f'B{output_bucket_cell}'] = timestamp
    ws[f'C{output_bucket_cell}'] = mass
    output_bucket_cell += 1
    wb.save(workbook_path)
    server.sendto(f"Logged Output Bucket At: {timestamp} \n Measured Mass of {mass}kg".ljust(100, '\0').encode(),
                  client_addr)


def handle_moisture(server, client_addr, workbook_path, data):
    """
    Collects and logs 10 moisture readings into the workbook.
    Sends a confirmation message to the client after all readings are logged.
    :param (socket.socket) server: The server which the client is listening to
    :param (tuple) client_addr: The IP and port number of the client
    :param (str) workbook_path: Path to the feeding log workbook
    :param (str) data: Request made by the client
    """
    wb = load_workbook(workbook_path)
    ws = wb[output_and_yield_worksheet]
    starting_cell = output_moisture_cell
    if re.search(r"^Input", data):
        starting_cell = input_moisture_cell
    for cell in range(10):
        moisture = get_data(server, f"Enter reading {cell + 1}: ".ljust(100, '\0'), client_addr, r"^\.\d{1,3}$",
                            f"Invalid Entry. Moisture must start with a decimal. " +
                            f"Please re-enter or send 'c' to cancel".ljust(100, '\0'))
        if moisture is None:
            return
        ws[f'B{starting_cell + cell}'] = float(moisture)
    wb.save(workbook_path)
    server.sendto(f"Moisture Readings Logged Successfully".ljust(100, '\0').encode(), client_addr)


def handle_water(server, client_addr, workbook_path, data):
    """
    Logs water flow rate or calculates total water used based on start and finish times.
    Sends a confirmation message to the client.
    :param (socket.socket) server: The server which the client is listening to
    :param (tuple) client_addr: The IP and port number of the client
    :param (str) workbook_path: Path to the feeding log workbook
    :param (str) data: Request made by the client
    :return:
    """
    global water_start_time
    global water_flow_rate
    global water_finish_time
    wb = load_workbook(workbook_path)
    if re.search(r"Flow", data):
        water_flow_rate = get_data(server, f"Enter Flow Rate (L/min)".ljust(100, '\0'), client_addr, r"^\d?\.\d{1,2}$",
                                   f"Invalid Flow Rate. Please re-enter or send 'c' to cancel".ljust(100, '\0'))
        water_flow_rate = float(water_flow_rate)

    t = time.localtime()
    t = t.tm_hour + t.tm_min / 60.0
    if re.search(r"Start", data):
        water_start_time = t
    else:
        water_finish_time = t
        ws = wb[output_and_yield_worksheet]
        ws[f'B{water_cell}'] = (water_finish_time - water_start_time) * water_flow_rate * 60
        wb.save(workbook_path)
    server.sendto(f"Water Reading Logged Successfully".ljust(100, '\0').encode(), client_addr)


def handle_datalog(server, client_addr, workbook_path, data):
    """
    Logs various observations or events, including smoke start/finish events.
    Allows clients to enter additional observations and comments.
    Sends a confirmation message upon successful logging.
    :param (socket.socket) server: The server which the client is listening to
    :param (tuple) client_addr: The IP and port number of the client
    :param (str) workbook_path: Path to the feeding log workbook
    :param (str) data: Request made by the client
    :return:
    """
    global datalog_cell
    wb = load_workbook(workbook_path)
    ws = wb[datalog_worksheet]
    timestamp = time.strftime("%I:%M:%S %p")
    comments = ''
    if re.search(hopper_smoke_start_pattern, data):
        observations = "Hopper Smoke Start (White)"
    elif re.search(hopper_smoke_finish_pattern, data):
        observations = "Hopper Smoke Finish (White)"
    elif re.search(chimney_smoke_start_pattern, data):
        observations = "Chimney Smoke Start (Black)"
    elif re.search(chimney_smoke_finish_pattern, data):
        observations = "Chimney Smoke Finish (Black)"
    else:
        observations = get_data(server, f"Enter Observation or 'c' to cancel".ljust(100, '\0'), client_addr)
        if observations is None:
            return
        comments = get_data(server, f"Enter Comments or 'c' to cancel".ljust(100, '\0'), client_addr)
        if comments is None:
            return
    ws[f'A{datalog_cell}'] = timestamp
    ws[f'B{datalog_cell}'] = observations
    ws[f'C{datalog_cell}'] = comments
    wb.save(workbook_path)
    server.sendto(f"{observations} Logged Successfully".ljust(100, '\0').encode(), client_addr)


def get_data(server, collection_message, client_addr, regex=r"^.*$",
             error_message=f"Invalid Entry. Please re-enter or send 'c' to cancel".ljust(100, '\0')):
    """
    Sends a prompt to the client, waits for a valid response matching the regex.
    If invalid, prompts again with an error message.
    Returns the valid input or None if canceled

    :param (socket.socket) server: The server which the client is listening to
    :param (str) collection_message: Message to the client to request an input
    :param (tuple) client_addr: The IP and port number of the client
    :param (str | optional) regex: The parameters that the input must fulfill
    :param (str | optional) error_message: Message to client if input does not satisfy the regex
    :return: The input collected from the client. If client cancels the transaction, returns None
    """
    server.sendto(collection_message.encode(), client_addr)
    while True:
        data, _ = server.recvfrom(1024)
        data = data.decode().strip()
        if re.match(regex, data):
            return data
        elif data == 'c':
            server.sendto(f"Entry cancelled".ljust(100, '\0').encode(), client_addr)
            return None
        else:
            server.sendto(error_message.encode(), client_addr)
